import io
import os
import re
import unicodedata
import uuid
from django.conf import settings
from django.http import FileResponse, HttpResponseNotFound, JsonResponse
from django.shortcuts import redirect, render
from openpyxl import load_workbook
import pandas as pd
from Db_handler.Database import EmailDatabase
from datetime import datetime
from django.core.files.storage import default_storage
from django.views.decorators.csrf import csrf_exempt
from django.core.files.base import ContentFile
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.cell import MergedCell
from ChatBot.Db_Handler.db import DatabaseManager





db = DatabaseManager()
def download_excel_view(request, id_msg):
    file_path = os.path.join(settings.MEDIA_ROOT, 'excel_responses', f'response_{id_msg}.xlsx')
    print("file_path", file_path)
    if os.path.exists(file_path):
        file_handle = open(file_path, 'rb')
        return FileResponse(file_handle, as_attachment=True, filename=f"response_{id_msg}.xlsx")
    else:
        print("Fichier non trouvé.")
        return HttpResponseNotFound("Fichier non trouvé.")
def clean_request(s: str):
    s = unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode('utf-8')
    

    # Suppression des espaces multiples
    s = re.sub(r"\s+", " ", s).strip()
    return s

def convert_excel(res, id_msg):
    file_path = os.path.join(settings.MEDIA_ROOT, 'excel_responses', f'report.xlsx')
    # Charger le modèle Excel
    wb = load_workbook(file_path)
    ws = wb.active

    # Mise à jour du titre en E1
    titre_cell = ws["E1"]
    titre_cell.value = ""
    titre_cell.font = Font(bold=False, size=16)
    titre_cell.alignment = Alignment(horizontal="center", vertical="center")

    border_style = Side(border_style="thin", color="000000")
    titre_cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Format de la date
    date_actuelle = datetime.now().strftime("%d %B %Y")  # Exemple : "10 Mai 2025"

    # Mettre à jour la date dans la feuille
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "Date d'édition du rapport:" in cell.value:
                cell.value = f"Date d'édition du rapport: {date_actuelle}"

    # Convertir les données en DataFrame
    df = pd.DataFrame(res)

    # Insertion du DataFrame à partir de la ligne 8
    start_row = 8
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == start_row:
                # Style de l'en-tête
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="dd33ff", end_color="dd33ff", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Bordures
            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Ajustement automatique de la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            if col_letter is None:
                col_letter = cell.column_letter

        if col_letter:
            ws.column_dimensions[col_letter].width = max_length + 2

    # Sauvegarde dans un fichier en mémoire
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Préparer pour l'enregistrement
    file_name = f"response_{id_msg}.xlsx"
    file_path = default_storage.save(
        f"excel_responses/{file_name}",
        ContentFile(output.read())
    )

    return file_path
def read_excel(id_msg):
    print("id_msg",id_msg)
    path_of_excel = "excel_responses/response_"
    file_path = f"{path_of_excel}{id_msg}.xlsx"
    if default_storage.exists(file_path):
        with default_storage.open(file_path, mode='rb') as excel_file:
            df = pd.read_excel(excel_file, engine='openpyxl',index_col=None,skiprows=7)
            df.columns = df.columns.astype(str)
            df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
            df = df.dropna(axis=1, how='all')
        return df

def send_request(payload: str):
    import requests
    import json

    url = "http://127.0.0.1:8081/process/"
    headers = {
        "Content-Type": "application/json"
    }
    data = {
        "request": clean_request(payload)
    }

    response = requests.post(url, headers=headers, data=json.dumps(data))

    print(response.status_code)
    try:
        
        return response.json()
    except json.JSONDecodeError:
        print("Réponse non JSON :", response.text)

# Exemple d'appel

# Create your views here.
def index(request):
    return render(request, 'chat.html')
def all_conversations(request):
    email_user = request.session.get('email_user')
    conversations = db.historique(user_id=email_user)
    return JsonResponse({"conversations": conversations})

def all_msg_by_conv(request, conv_id):
    email_user = request.session.get('email_user')
    type_res = "str"
    message = ""
    if conv_id:
        messages = db.get_all_msg_con(conv_id, user_id=email_user)
        for message in messages:
            if message.get("res_user").endswith(".xlsx"):
                 print(True)
                 message_x = read_excel(message.get("msg_id"))
                 message["message"] = message_x.to_dict(orient="records")
                 message["type_res"] = "list"
            else :
                message["type_res"] = "str"

                
        return JsonResponse({"messages": messages})
    return JsonResponse({"error": "conv_id missing"}, status=400)


@csrf_exempt 
def create_conversation(request): 
    if request.method != "POST":
        return JsonResponse({"error": "Only POST method is allowed"}, status=405)

    msg = request.POST.get('message')
    conv_id_from_post = request.POST.get('conv_id', None) 

    current_date = datetime.now()
    format_date = current_date.strftime("%Y-%m-%d %H:%M:%S")
    
    user_id = request.session.get('email_user')

    if not msg:
        return JsonResponse({"error": "No message provided"}, status=400)

    final_conv_id = None 

  
    if conv_id_from_post:
        print(f"Received existing conv_id: {conv_id_from_post}") 
      
        msg_id = db.insert_chat(msg, format_date, user_id, conv_id_from_post)
        print(msg_id)
        final_conv_id = conv_id_from_post 
    else:
        print("No conv_id received, creating new conversation.") 
        new_conv_id = str(uuid.uuid4())
        msg_id = db.insert_chat(msg, format_date, user_id, new_conv_id)
        final_conv_id = new_conv_id 

    
    bot_response = {
        "msg_id": msg_id,
        "role": "bot",
        "content": msg, 
        "timestamp": format_date, 
    }

    return JsonResponse({
        "status": "inserted",
        "conv_id": final_conv_id, 
        "message": bot_response
    })
@csrf_exempt
def update_conversation(request, msg_user, id_msg):
    print(f"update_conversation called with msg_user='{msg_user}', id_msg='{id_msg}'")

    response = send_request(clean_request(msg_user))
    res= response.get("response", "")
    req_sql = response.get("sql_query", "")
    print("le type de res est ",type(res))
    date_res = datetime.now()
    format_date = date_res.strftime("%Y-%m-%d %H:%M:%S")

    

    if response:
        ct_msg = response.get("category", None)
        if type(res) == str:
              
                    type_res = "text"
                    success = db.insert_res_msg(res, format_date, id_msg,req_sql,ct_msg)
                    print("Insert success:", success)
        else:
                   
                    type_res = "list"
                    file_path = convert_excel(res,id_msg)
                    success = db.insert_res_msg(file_path, format_date, id_msg,req_sql,ct_msg)
                    print("Insert success:", success)



        if success:
            return JsonResponse({
                "status": "updated",
                "response": res,
                "format_date": format_date,
                "id_msg": id_msg,
                "type_res": type_res
            })
        else:
            return JsonResponse({"error": "DB insert failed"}, status=400)
    else:
        return JsonResponse({"error": "Ollama response empty"}, status=400)



 